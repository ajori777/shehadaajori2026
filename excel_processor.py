import openpyxl
import re
import pandas as pd
from collections import Counter
import os

class ExcelProcessor:
    def __init__(self):
        self.all_data = [] # List of dicts, each representing a sheet/class
        self.school_info_list = [] # List to aggregate school info for voting

    def clean_text(self, text):
        if text is None:
            return ""
        return str(text).strip().replace('\n', ' ')

    def extract_from_regex(self, text, patterns):
        results = {}
        for key, pattern in patterns.items():
            match = re.search(pattern, text)
            if match:
                results[key] = match.group(1).strip()
        return results

    def get_merged_value(self, sheet, row, col):
        cell = sheet.cell(row=row, column=col)
        for merged_range in sheet.merged_cells.ranges:
            if cell.coordinate in merged_range:
                return sheet.cell(row=merged_range.min_row, column=merged_range.min_col).value
        return cell.value

    def process_file(self, file_path):
        wb = openpyxl.load_workbook(file_path, data_only=True)
        file_results = []
        
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            # Basic validation: check if it looks like a student sheet
            # A1 usually contains table title
            title_cell = sheet.cell(row=1, column=1).value
            if not title_cell or ("جدول العلامات" not in str(title_cell) and "الشهادة المدرسية" not in str(title_cell)):
                 # Try searching a bit more or just continue
                 if not any("علامات" in str(sheet.cell(row=r, column=1).value) for r in range(1, 5)):
                     continue

            data = self.parse_sheet(sheet, file_path)
            if data:
                file_results.append(data)
                self.school_info_list.append(data['school_info'])
        
        return file_results

    def normalize_arabic(self, text):
        if not text: return ""
        text = str(text).strip()
        # Normalize Alif Hamza
        text = re.sub(r'[أإآ]', 'ا', text)
        # Normalize Teh Marbuta
        text = re.sub(r'ة', 'ه', text)
        # Remove extra spaces
        text = re.sub(r'\s+', ' ', text)
        return text

    def parse_sheet(self, sheet, file_path):
        # 1. School Metadata (Rows 1-6)
        school_info = {
            'school_name': '',
            'national_id': '',
            'directorate': '',
            'district': '',
            'town': '',
            'grade': '',
            'section': '',
            'year': '',
            'principal_name': ''
        }
        
        # Regex patterns for extraction
        patterns = {
            'directorate': r'مديرية التربية والتعليم:\s*(.*)',
            'town': r'البلدة:\s*(.*?)(?:\s+المدرسة|$)',
            'school_name': r'المدرسة:\s*(.*?)(?:\s*-|$)',
            'national_id': r'-\s*(\d+)',
            'district': r'اللواء:\s*(.*?)(?:\s+الشعبة|$)',
            'section': r'الشعبة:\s*(.*)',
            'grade': r'للصف\s+(.*?)(?:\s+مسار|$|\s+للعام)',
            'year': r'للعام الدراسي\s*\((.*?)\)'
        }

        for r in range(1, 7):
            for c in range(1, 60): # Scan broad range
                val = self.clean_text(sheet.cell(row=r, column=c).value)
                if val:
                    extracted = self.extract_from_regex(val, patterns)
                    for k, v in extracted.items():
                        if v: school_info[k] = v

        # 2. Identify Subject Columns & Special Columns
        subjects = []
        special_cols = {'absence': None, 'total': None, 'avg_pct': None, 'result': None, 'field': None}

        # Find the anchor column ('اسم الطالب' or 'الرقم الوطني' or 'الاسم')
        anchor_col = 2 # Default to B
        for r in range(7, 10):
            for c in range(1, 15):
                cell_val = self.clean_text(sheet.cell(row=r, column=c).value)
                if any(x in cell_val for x in ["اسم الطالب", "الرقم الوطني"]) or cell_val == "الاسم":
                    anchor_col = c
                    break
        
        # Specifically look for 'Field/Stream'
        for c in range(anchor_col, anchor_col + 10):
            v7 = self.clean_text(sheet.cell(row=7, column=c).value)
            v8 = self.clean_text(sheet.cell(row=8, column=c).value)
            if any(x in v7 or x in v8 for x in ["فرع", "مسار", "اسم الحقل"]):
                special_cols['field'] = c
                break
        
        # Find the first score column in Row 8
        first_score_col = None
        for c in range(anchor_col + 1, sheet.max_column + 1):
            v8 = self.clean_text(sheet.cell(row=8, column=c).value)
            if any(x in v8 for x in ["الفصل الأول", "الفصل الثاني", "المعدل"]):
                first_score_col = c
                break
        
        start_col = first_score_col if first_score_col else (anchor_col + 1)

        # Scan Row 7 for subjects, starting after personal info/field
        current_subject = None
        for c in range(start_col, sheet.max_column + 1):
            val = self.get_merged_value(sheet, 7, c)
            sub_val = self.clean_text(sheet.cell(row=8, column=c).value)
            
            if val and val != current_subject:
                # Basic check: is it actually a subject or a final metadata column?
                if not any(x in str(val) for x in ["المجموع", "المعدل", "النتيجة", "الغياب", "تعبأ المعلومات"]):
                    norm_val = self.normalize_arabic(val)
                    if "التربية الاسلامية" in norm_val:
                        val = "التربية الإسلامية"
                    current_subject = val
                    subjects.append({'name': val, 'start_col': c, 'sub_cols': {}})
                else:
                    current_subject = None
            
            if current_subject:
                if "الفصل الأول" in sub_val: subjects[-1]['sub_cols']['t1'] = c
                elif "الفصل الثاني" in sub_val: subjects[-1]['sub_cols']['t2'] = c
                elif "المعدل" in sub_val: subjects[-1]['sub_cols']['avg'] = c

            if "عدد أيام غياب" in sub_val: special_cols['absence'] = c
            elif "المجموع العام" in sub_val: special_cols['total'] = c
            elif "المعدل العام المئوي" in sub_val: special_cols['avg_pct'] = c
            elif "النتيجة السنوية" in sub_val: special_cols['result'] = c

        subjects = [s for s in subjects if s['sub_cols']]

        # 3. Teacher and Principal (Bottom of sheet)
        teacher_name = ""
        principal_name = ""
        teacher_label_col = None
        principal_label_col = None
        f2_row = None
        f1_row = None

        for r in range(sheet.max_row, max(1, sheet.max_row - 40), -1):
            for c in range(1, 60):
                val = self.clean_text(sheet.cell(row=r, column=c).value)
                if "مربي الصف" in val and not teacher_label_col:
                    teacher_label_col = c
                if "مدير المدرسة" in val and not principal_label_col:
                    principal_label_col = c
                if "الفصل الثاني" in val and not f2_row:
                    f2_row = r
                if "الفصل الأول" in val and not f1_row:
                    f1_row = r
            if teacher_label_col and principal_label_col and f2_row:
                break

        def find_name_near(label_col, target_row, hint_col=None):
            if not target_row: return ""
            search_cols = []
            if hint_col: search_cols.append(hint_col)
            if label_col:
                search_cols.extend([label_col + 1, label_col + 2, label_col])
            search_cols = list(dict.fromkeys([c for c in search_cols if c and c > 0]))
            skip_terms = ["اسم وتوقيع", "مربي الصف", "مدير المدرسة", "التاريخ", "الفصل", "الخاتم"]
            for col in search_cols:
                val = self.clean_text(sheet.cell(row=target_row, column=col).value)
                if val:
                    is_label = any(term in val for term in skip_terms)
                    if not is_label and len(val) > 2:
                         return val
            return ""

        teacher_name = find_name_near(teacher_label_col, f2_row, hint_col=11)
        if not teacher_name: teacher_name = find_name_near(teacher_label_col, f1_row, hint_col=11)
        principal_name = find_name_near(principal_label_col, f2_row, hint_col=27)
        if not principal_name: principal_name = find_name_near(principal_label_col, f1_row, hint_col=27)
        school_info['principal_name'] = principal_name

        # 4. Student Data (Row 10 onwards)
        students = []
        max_grades = {}
        min_grades = {}
        
        first_student_row = 10
        for r in range(8, sheet.max_row + 1):
            serial = sheet.cell(row=r, column=1).value
            if serial and str(serial).isdigit():
                first_student_row = r
                break
        
        max_grade_row = first_student_row - 1
        for sub in subjects:
            s_name = sub['name']
            for term_key in ['t1', 't2', 'avg']:
                col = sub['sub_cols'].get(term_key)
                if col:
                    max_val = sheet.cell(row=max_grade_row, column=col).value
                    try:
                        max_val = int(float(max_val)) if max_val is not None else 100
                    except:
                        max_val = 100
                    max_grades[(s_name, term_key)] = max_val
                    min_grades[(s_name, term_key)] = max_val // 2

        for r in range(first_student_row, sheet.max_row + 1):
            serial = sheet.cell(row=r, column=1).value
            if not serial or not str(serial).isdigit():
                if len(students) > 0: break
                continue
            
            name = self.clean_text(sheet.cell(row=r, column=2).value)
            nationality = self.clean_text(sheet.cell(row=r, column=3).value)
            pob = self.clean_text(sheet.cell(row=r, column=4).value)
            
            day = sheet.cell(row=r, column=5).value
            month = sheet.cell(row=r, column=6).value
            year = sheet.cell(row=r, column=7).value
            dob = f"{year}-{month}-{day}" if year and month and day else ""
            
            student_grades = {}
            for sub in subjects:
                s_name = sub['name']
                t1_col = sub['sub_cols'].get('t1')
                t2_col = sub['sub_cols'].get('t2')
                avg_col = sub['sub_cols'].get('avg')
                
                student_grades[f"{s_name}_ف1"] = sheet.cell(row=r, column=t1_col).value if t1_col else ""
                student_grades[f"{s_name}_ف2"] = sheet.cell(row=r, column=t2_col).value if t2_col else ""
                student_grades[f"{s_name}_معدل"] = sheet.cell(row=r, column=avg_col).value if avg_col else ""
                
                student_grades[f"{s_name}_ف1_عظمى"] = max_grades.get((s_name, 't1'), 100)
                student_grades[f"{s_name}_ف1_صغرى"] = min_grades.get((s_name, 't1'), 50)
                student_grades[f"{s_name}_ف2_عظمى"] = max_grades.get((s_name, 't2'), 100)
                student_grades[f"{s_name}_ف2_صغرى"] = min_grades.get((s_name, 't2'), 50)
                student_grades[f"{s_name}_معدل_عظمى"] = max_grades.get((s_name, 'avg'), 100)
                student_grades[f"{s_name}_معدل_صغرى"] = min_grades.get((s_name, 'avg'), 50)
            
            student_grades['الغياب'] = sheet.cell(row=r, column=special_cols['absence']).value if special_cols['absence'] else ""
            student_grades['المجموع'] = sheet.cell(row=r, column=special_cols['total']).value if special_cols['total'] else ""
            student_grades['المعدل_المئوي'] = sheet.cell(row=r, column=special_cols['avg_pct']).value if special_cols['avg_pct'] else ""
            student_grades['النتيجة'] = sheet.cell(row=r, column=special_cols['result']).value if special_cols['result'] else ""
            student_grades['field'] = sheet.cell(row=r, column=special_cols['field']).value if special_cols.get('field') else ""

            students.append({
                'serial': serial,
                'name': name,
                'nationality': nationality,
                'pob': pob,
                'dob': dob,
                **student_grades
            })

        return {
            'file_path': file_path,
            'sheet_name': sheet.title,
            'school_info': school_info,
            'teacher_name': teacher_name,
            'principal_name': principal_name,
            'students': students,
            'subjects': [s['name'] for s in subjects],
            'max_grades': max_grades,
            'min_grades': min_grades
        }

    def get_aggregated_school_info(self):
        if not self.school_info_list: return {}
        agg = {}
        keys = self.school_info_list[0].keys()
        for k in keys:
            vals = [info[k] for info in self.school_info_list if info[k]]
            if vals:
                agg[k] = Counter(vals).most_common(1)[0][0]
            else:
                agg[k] = ""
        return agg

    def export_to_excel(self, class_data, output_path):
        df = pd.DataFrame(class_data['students'])
        df.to_excel(output_path, index=False)
