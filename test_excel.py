from excel_processor import ExcelProcessor
import os
import openpyxl

def test_extraction():
    processor = ExcelProcessor()
    test_file = 'EP_StudentCertificate (1)-1.xlsx'
    if not os.path.exists(test_file):
        print(f"File {test_file} not found.")
        return

    results = processor.process_file(test_file)
    print(f"Processed {len(results)} sheets.")
    
    for res in results:
        print(f"\n--- Sheet: {res['sheet_name']} ---")
        # Debug bottom rows
        wb = openpyxl.load_workbook(test_file, data_only=True)
        sheet = wb[res['sheet_name']]
        print("Bottom rows debug:")
        for r in range(sheet.max_row, max(1, sheet.max_row - 30), -1):
            row_vals = [sheet.cell(row=r, column=c).value for c in range(1, 40)]
            if any(row_vals):
                print(f"Row {r}: {row_vals}")

        print(f"School Info: {res['school_info']}")
        print(f"Teacher: {res['teacher_name']}, Principal: {res['principal_name']}")
        print(f"Number of Students: {len(res['students'])}")
        if res['students']:
            print(f"First Student: {res['students'][0]['name']} - DOB: {res['students'][0]['dob']}")
            # Print subjects
            print(f"Subjects: {res['subjects']}")
            # Sample grades for first student
            sample_sub = res['subjects'][0]
            print(f"Grades for {sample_sub}: T1={res['students'][0].get(f'{sample_sub}_ف1')}, T2={res['students'][0].get(f'{sample_sub}_ف2')}")

if __name__ == "__main__":
    test_extraction()
